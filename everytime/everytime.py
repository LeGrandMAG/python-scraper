from selenium import webdriver
from bs4 import BeautifulSoup
from time import sleep
from openpyxl import Workbook,load_workbook

chrome_options = webdriver.ChromeOptions()
prefs = {"profile.default_content_setting_values.notifications" : 2}
chrome_options.add_experimental_option("prefs",prefs)


# Chrome의 경우 | 아까 받은 chromedriver의 위치를 지정해준다.
driver = webdriver.Chrome('C:/Users/maglo/OneDrive/Documents/chromedriver.exe',options=chrome_options)

# url에 접근한다.
driver.get('https://everytime.kr/login')
# 암묵적으로 웹 자원 로드를 위해 5초까지 기다려 준다.
driver.implicitly_wait(3)

# 아이디/비밀번호를 입력해준다.
driver.find_element_by_name('userid').send_keys('201800071')
driver.find_element_by_name('password').send_keys('magl0ire596')

sleep(2)

# 로그인 버튼을 눌러주자.
driver.find_element_by_xpath('//*[@id="container"]/form/p[3]/input').click()

driver.get('https://everytime.kr/timetable')

#수업 목록에서 검색 클릭
driver.find_element_by_xpath('//*[@id="container"]/ul/li[1]').click()

#팝업창 닫기
sleep(2)
#driver.find_element_by_xpath('//*[@id="sheet"]/ul/li[3]/a').click()

pre_count = 0
#스크롤 맨아래로 내리기
while True:
    #tr요소 접근
    element = driver.find_elements_by_css_selector("#subjects > div.list > table > tbody > tr")

    # tr 마지막 요소 접근
    result = element[-1]
    #마지막요소에 focus주기
    driver.execute_script('arguments[0].scrollIntoView(true);',result)
    sleep(2)

    #현재 접근한 요소의 갯수
    current_count = len(element)
    if pre_count == current_count:
        break
    #같지않다면
    pre_count = current_count


html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')

trs = soup.select('#subjects > div.list > table > tbody > tr')

results = []

for tr in range(len(trs)):
    result= { 
        'CoursePl': [],
        'Category': [],
        'Year': [],
        'CourseID': [],
        'CourseName':[],
        'Credit': [],
        'Duration': [],
        'Professor': [],
        'Schedule': [],
        'CurrentStudents':[],
        'TotalStudents':[],
        'Note':[],
    }


    
    tds = trs[tr].select('#subjects > div.list > table > tbody > tr > td')
    result['CoursePl'].append(tds[0].text) #Course Plan
    result['Category'].append(tds[1].text) #Category
    result['Year'].append(tds[2].text) #Year
    result['CourseID'].append(tds[3].text) #Course Code
    result['CourseName'].append(tds[4].text) #Course Name
    result['Credit'].append(tds[5].text) #Credit
    result['Duration'].append(tds[6].text) #Duration
    result['Professor'].append(tds[7].text) #Professor Name
    result['Schedule'].append(tds[8].text) #Schedule
    result['CurrentStudents'].append(tds[10].text) #CurrentStudents
    result['TotalStudents'].append(tds[11].text) #TotalStudents
    result['Note'].append(tds[12].text) #Note
    results.append(result)
#값이 들어있다면!
if results:
    print("성공!!")

"""
excel_column = 9
write_wb = Workbook()
write_ws = write_wb.create_sheet('result.xls')
for data in results:
    write_ws = write_wb.active
    write_ws.append(data)
write_wb.save('result.csv') """